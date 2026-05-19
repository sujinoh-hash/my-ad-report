"""
GA4 TSV → Excel 변환 스크립트
사용법: python ga4_tsv_to_excel.py [TSV파일경로] [출력파일명(선택)]
예시:  python ga4_tsv_to_excel.py ga4_report.tsv
       python ga4_tsv_to_excel.py ga4_report.tsv 결과.xlsx
"""

import sys
import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# GA4 TSV 파싱 (헤더 자동 감지)
# ─────────────────────────────────────────────
def read_ga4_tsv(tsv_path: str) -> pd.DataFrame:
    """
    GA4 TSV 파일을 읽는다.
    - 상단에 메타 정보(보고서 기간 등) 행이 있어도 자동 건너뜀
    - 탭 구분자 기준으로 파싱
    - 하단 합계 행("합계", "총계", "Total") 제거
    """
    header_row = None

    with open(tsv_path, encoding="utf-8-sig") as f:
        lines = f.readlines()

    for i, line in enumerate(lines):
        cols = [c.strip() for c in line.strip().split("\t")]
        # 컬럼 헤더 행: 날짜 또는 세션 관련 컬럼 포함 여부로 판별
        if any(k in " ".join(cols) for k in ["날짜", "세션", "Date", "Session", "Campaign"]):
            header_row = i
            break

    if header_row is None:
        raise ValueError("헤더 행을 찾을 수 없습니다. TSV 파일 구조를 확인하세요.")

    df = pd.read_csv(tsv_path, sep="\t", header=header_row, encoding="utf-8-sig")

    # 하단 합계 행 제거
    if len(df) > 0:
        first_col = df.columns[0]
        df = df[~df[first_col].astype(str).str.contains(r"합계|총계|Total|Grand Total", na=False)]

    # 빈 행 제거
    df.dropna(how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)

    return df


# ─────────────────────────────────────────────
# 컬럼 자동 감지 & 한글 컬럼명 정규화
# ─────────────────────────────────────────────
COLUMN_MAP = {
    # 날짜
    "날짜": "날짜",
    "date": "날짜",
    # 소스
    "세션 소스": "세션소스",
    "세션소스": "세션소스",
    "session source": "세션소스",
    # 캠페인
    "세션 캠페인": "세션캠페인",
    "세션캠페인": "세션캠페인",
    "session campaign": "세션캠페인",
    "campaign": "세션캠페인",
    # 광고콘텐츠
    "세션 광고 콘텐츠": "세션광고콘텐츠",
    "세션광고콘텐츠": "세션광고콘텐츠",
    "세션 수동 광고 콘텐츠": "세션광고콘텐츠",
    "session ad content": "세션광고콘텐츠",
    # 검색어
    "세션 검색어": "세션검색어",
    "세션검색어": "세션검색어",
    "session search term": "세션검색어",
    # 방문수
    "방문수": "방문수",
    "users": "방문수",
    "총 사용자": "방문수",
    # 참여세션수
    "참여 세션수": "참여세션수",
    "참여세션수": "참여세션수",
    "engaged sessions": "참여세션수",
    # 세션수
    "세션수": "세션수",
    "세션": "세션수",
    "sessions": "세션수",
    # 장바구니
    "장바구니에 추가": "장바구니",
    "장바구니": "장바구니",
    "add to carts": "장바구니",
    "add to cart": "장바구니",
    # 구매
    "구매": "구매",
    "purchases": "구매",
    "transactions": "구매",
    # 수익
    "총 수익": "총수익",
    "총수익": "총수익",
    "revenue": "총수익",
    "total revenue": "총수익",
    "구매 수익": "총수익",
}

def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    ren = {}
    for col in df.columns:
        key = col.strip().lower()
        for pattern, mapped in COLUMN_MAP.items():
            if key == pattern.lower():
                ren[col] = mapped
                break
    df.rename(columns=ren, inplace=True)
    return df


# ─────────────────────────────────────────────
# 캠페인명 변환 (어도비와 동일 규칙)
# ─────────────────────────────────────────────
def normalize_campaign_key(raw: str) -> str:
    return re.sub(r"\s+", "-", raw.strip().lower())

def build_campaign_name(cid: str) -> str:
    if pd.isna(cid) or str(cid).strip() in ("", "(not set)"):
        return "Unknown"

    raw = str(cid).strip()
    low = raw.lower()
    parts = raw.split("_")
    prefix = parts[0].lower()

    if raw.startswith("dm-"):
        return raw

    if "_adef-" in low or "_br_" in low:
        return "Unknown"

    if prefix == "dsp":
        medium = parts[1].lower() if len(parts) > 1 else ""
        seg7   = parts[7].lower() if len(parts) > 7 else ""
        raw_key = parts[8] if len(parts) > 8 else ""

        if medium == "google":
            if not seg7.startswith("prospecting"):
                return "Unknown"
            if "demo-women" in seg7:   pmax = "PmaxW"
            elif "demo-men" in seg7:   pmax = "PmaxM"
            else:                      pmax = "PmaxC"
            return f"dm-prospecting-{pmax}-alwayson-na-na"

        if medium == "criteo":
            seg6 = parts[6].lower() if len(parts) > 6 else ""
            if "retargeting" in seg6 or "retargeting" in seg7:
                return "dm-retargeting-criteo-alwayson-na-na"
            return "Unknown"

        if medium in ("kakao", "kakaoda"):
            seg6 = parts[6].lower() if len(parts) > 6 else ""
            c_key = normalize_campaign_key(raw_key) if raw_key else ""
            if "retargeting" in seg6 or "retargeting" in seg7:
                return f"dm-retargeting-kakao-{c_key}-na-na" if c_key else "dm-retargeting-kakao-alwayson-na-na"
            if "prospecting" in seg6 or "prospecting" in seg7:
                return f"dm-prospecting-kakao-{c_key}-na-na" if c_key else "dm-prospecting-kakao-alwayson-na-na"
            return "Unknown"

        if medium == "meta":
            seg6 = parts[6].lower() if len(parts) > 6 else ""
            c_key = normalize_campaign_key(raw_key) if raw_key else ""
            if "retargeting" in seg6 or "retargeting" in seg7:
                return f"dm-retargeting-meta-{c_key}-na-na" if c_key else "dm-retargeting-meta-alwayson-na-na"
            if "prospecting" in seg6 or "prospecting" in seg7:
                return f"dm-prospecting-meta-{c_key}-na-na" if c_key else "dm-prospecting-meta-alwayson-na-na"
            return "Unknown"

        return "Unknown"

    # SEM (google, naver 등)
    if prefix in ("sem", "sa"):
        medium = parts[1].lower() if len(parts) > 1 else ""
        seg5   = parts[5].lower() if len(parts) > 5 else ""
        seg6   = parts[6].lower() if len(parts) > 6 else ""

        if medium == "google":
            if "brand" in seg5 or "brand" in seg6:
                return "dm-brand-googlepcmo-brand-na-na"
            elif "nonbrand" in seg5 or "non-brand" in seg5:
                if "product" in seg6:
                    return "dm-nonbrand-googlepcmo-product-na-na"
                return "dm-nonbrand-googlepcmo-general-na-na"
            return "Unknown"

        if medium == "naver":
            if "brand" in seg5 or "brand" in seg6:
                return "dm-brand-naversa-brand-na-na"
            elif "nonbrand" in seg5 or "non-brand" in seg5:
                return "dm-nonbrand-naversa-general-na-na"
            return "Unknown"

    return "Unknown"


# ─────────────────────────────────────────────
# 숫자 컬럼 타입 변환
# ─────────────────────────────────────────────
NUMERIC_COLS = ["방문수", "참여세션수", "세션수", "장바구니", "구매", "총수익"]

def cast_numerics(df: pd.DataFrame) -> pd.DataFrame:
    for col in NUMERIC_COLS:
        if col in df.columns:
            df[col] = (
                df[col].astype(str)
                       .str.replace(",", "", regex=False)
                       .str.replace("₩", "", regex=False)
                       .str.strip()
            )
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


# ─────────────────────────────────────────────
# Excel 저장 (스타일 포함)
# ─────────────────────────────────────────────
HEADER_BG   = "4472C4"
HEADER_FONT = "FFFFFF"
ROW_EVEN_BG = "DCE6F1"

def save_to_excel(df: pd.DataFrame, out_path: str):
    df.to_excel(out_path, index=False, engine="openpyxl")

    wb = load_workbook(out_path)
    ws = wb.active
    ws.title = "GA4 데이터"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 헤더 스타일
    for cell in ws[1]:
        cell.font      = Font(name="Arial", bold=True, color=HEADER_FONT, size=10)
        cell.fill      = PatternFill("solid", start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    # 데이터 행 스타일
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        bg = ROW_EVEN_BG if row_idx % 2 == 0 else "FFFFFF"
        for cell in row:
            cell.font   = Font(name="Arial", size=10)
            cell.fill   = PatternFill("solid", start_color=bg)
            cell.border = border
            col_name    = ws.cell(1, cell.column).value or ""

            # 숫자 컬럼 서식
            if col_name in ("총수익",):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif col_name in ("방문수", "참여세션수", "세션수", "장바구니", "구매"):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif col_name == "날짜":
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

    # 컬럼 너비 자동 조정
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # 행 높이
    ws.row_dimensions[1].height = 28
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 18

    # 틀 고정 (헤더)
    ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"✅ 저장 완료: {out_path}  ({ws.max_row - 1}행)")


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("사용법: python ga4_tsv_to_excel.py <TSV파일경로> [출력파일명.xlsx]")
        sys.exit(1)

    tsv_path = sys.argv[1]
    if not os.path.exists(tsv_path):
        print(f"파일 없음: {tsv_path}")
        sys.exit(1)

    out_path = sys.argv[2] if len(sys.argv) > 2 else os.path.splitext(tsv_path)[0] + ".xlsx"

    print(f"📂 읽는 중: {tsv_path}")
    df = read_ga4_tsv(tsv_path)
    print(f"   원본 행수: {len(df)}, 컬럼수: {len(df.columns)}")

    df = normalize_columns(df)
    df = cast_numerics(df)

    # 캠페인명 변환 컬럼 추가
    if "세션캠페인" in df.columns:
        df.insert(
            df.columns.get_loc("세션캠페인") + 1,
            "캠페인명",
            df["세션캠페인"].apply(build_campaign_name)
        )
        unknown_cnt = (df["캠페인명"] == "Unknown").sum()
        print(f"   캠페인명 변환 완료 (Unknown: {unknown_cnt}행 / {len(df)}행)")

    save_to_excel(df, out_path)


if __name__ == "__main__":
    main()
